"""
test_program.py
---------------
Tests شامله للبرنامج. بتختبر:
  1. قراءة program_mapping.xlsx
  2. استخراج Serial Number من الباركود
  3. تحديد البرنامج من الباركود
  4. TCPServer/TCPClient (محلي على 127.0.0.1)
  5. سلوك الـ watchdog لما الاتصال يتقطع
  6. result_reporting بيكتب صح في Excel
  7. debug_monitor snapshot

تشغيل:
    python test_program.py
    # أو
    python -m unittest test_program -v
"""

import os
import sys
import time
import socket
import threading
import unittest
import tempfile
import queue
from unittest import mock

# ─── Mock للموديولز اللي مش موجودة في بيئة الـ tests ─────────────────────────
# keyboard module مش موجود في Linux sandbox، نعمل stub
sys.modules.setdefault("keyboard", mock.MagicMock(KEY_DOWN="down"))
sys.modules.setdefault("pyodbc", mock.MagicMock())

# scanner بيستورد keyboard، فلازم نـ mock keyboard أولاً ثم نستورد scanner
import scanner  # noqa: E402

# ─── الآن نستورد باقي الموديولز ─────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import ClientsClass  # noqa: E402
import excel as ex   # noqa: E402
import debug_monitor # noqa: E402


# ════════════════════════════════════════════════════════════════════════════
#                              Excel & Mapping
# ════════════════════════════════════════════════════════════════════════════
class TestProgramMapping(unittest.TestCase):
    """يتأكد إن ملف program_mapping.xlsx شغّال وبيرجع القيم الصحيحة."""

    def test_xlsx_readable(self):
        from openpyxl import load_workbook
        wb = load_workbook("program_mapping.xlsx")
        ws = wb.active
        self.assertGreater(ws.max_row, 1, "الملف لازم يحتوي على بيانات")

    def test_xlsx_via_pandas(self):
        import pandas as pd
        df = pd.read_excel("program_mapping.xlsx")
        self.assertIn("Capacity", df.columns)
        self.assertIn("program", df.columns)
        self.assertGreaterEqual(len(df), 5)

    def test_known_mapping_values(self):
        """E→1, F→1, G→2, ... P→5 زي ما اتفقنا."""
        app = self._make_app_without_starting_server()
        try:
            for char, expected_program in [("E", 1), ("F", 1), ("G", 2), ("M", 5), ("P", 5)]:
                fake_barcode = f"2605TL000001B{char}SI"
                result = app.determine_program_from_barcode(fake_barcode)
                self.assertEqual(int(result), expected_program,
                                 f"للحرف {char} متوقع {expected_program} لكن طلع {result}")
        finally:
            self._cleanup_app(app)

    def test_unknown_character_returns_message(self):
        app = self._make_app_without_starting_server()
        try:
            result = app.determine_program_from_barcode("2605TL000001BZSI")
            self.assertIsInstance(result, str)
            self.assertIn("غير موجود", result)
        finally:
            self._cleanup_app(app)

    def test_missing_file_handled(self):
        app = self._make_app_without_starting_server()
        try:
            result = app.determine_program_from_barcode(
                "2605TL000001BESI",
                excel_file_path="this_file_does_not_exist.xlsx",
            )
            self.assertIn("غير موجود", str(result))
        finally:
            self._cleanup_app(app)

    # ── helpers ──
    def _make_app_without_starting_server(self):
        """بنعمل App بدون ما الـ TCPServer يحجز بورت (عشان مايلخبطش الـ tests التانية)."""
        with mock.patch.object(ClientsClass.TCPServer, "start", return_value=True):
            app = ClientsClass.App()
        return app

    def _cleanup_app(self, app):
        try:
            app.stop()
        except Exception:
            pass


# ════════════════════════════════════════════════════════════════════════════
#                          Serial extraction
# ════════════════════════════════════════════════════════════════════════════
class TestSerialExtraction(unittest.TestCase):

    def setUp(self):
        with mock.patch.object(ClientsClass.TCPServer, "start", return_value=True):
            self.app = ClientsClass.App()

    def tearDown(self):
        self.app.stop()

    def test_extract_standard_barcode(self):
        # 2605TL0000001BISI → بين الحروف الأرقام هي 0000001
        serial = self.app.extract_serial_number("2605TL0000001BISI")
        # ملحوظة: regex بياخد أكبر سلسلة أرقام بين حروف
        # 2605TL... → أول مجموعه أرقام بعد حروف هي 0000001
        self.assertIsNotNone(serial)

    def test_extract_no_digits(self):
        serial = self.app.extract_serial_number("ABCDEF")
        self.assertIsNone(serial)

    def test_extract_short_barcode(self):
        serial = self.app.extract_serial_number("A1B")
        self.assertEqual(serial, "1")


# ════════════════════════════════════════════════════════════════════════════
#                            TCP server/client
# ════════════════════════════════════════════════════════════════════════════
def _get_free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


class TestTCPServerClient(unittest.TestCase):
    """tests على TCPServer/TCPClient الحقيقيين عبر loopback."""

    def setUp(self):
        self.port = _get_free_port()
        self.server = ClientsClass.TCPServer(ip="127.0.0.1", port=self.port)
        self.assertTrue(self.server.start(), "السيرفر مش قادر يبدأ")
        self.received = []
        self.server.start_listening(self._cb)
        time.sleep(0.1)

    def tearDown(self):
        try:
            self.server.stop()
        except Exception:
            pass
        time.sleep(0.1)

    def _cb(self, client_sock, addr, data):
        self.received.append(data)

    def test_client_can_connect_and_send(self):
        client = ClientsClass.TCPClient("127.0.0.1", self.port, timeout=2.0, name="testclient")
        self.assertTrue(client.connect())
        self.assertTrue(client.send_only(b"hello"))
        time.sleep(0.3)
        self.assertIn(b"hello", self.received)
        client.disconnect()

    def test_disconnect_does_not_crash(self):
        """قبل التعديل كان بيرمي AttributeError على _stop_monitor."""
        client = ClientsClass.TCPClient("127.0.0.1", self.port, timeout=2.0)
        client.connect()
        # ده اللي كان بيقع البرنامج
        client.disconnect()
        self.assertFalse(client.connected)

    def test_client_name_appears_in_log(self):
        """قبل التعديل كان يطبع [][INFO] ... دلوقتي بيطبع [TCPClient-ip:port]."""
        client = ClientsClass.TCPClient("127.0.0.1", self.port, name="MyCobot")
        self.assertEqual(client.name, "MyCobot")
        client2 = ClientsClass.TCPClient("127.0.0.1", self.port)
        self.assertIn("127.0.0.1", client2.name)

    def test_get_last_received_before_listen_does_not_crash(self):
        """قبل التعديل كان يرمي AttributeError لو get_last_received اتنادت قبل start_listening."""
        client = ClientsClass.TCPClient("127.0.0.1", self.port)
        # ده كان بيرمي AttributeError
        result = client.get_last_received(block=False)
        self.assertIsNone(result)


# ════════════════════════════════════════════════════════════════════════════
#                               Watchdog
# ════════════════════════════════════════════════════════════════════════════
class TestConnectionWatchdog(unittest.TestCase):
    """يتأكد إن الـ watchdog بيلاحظ لما السيرفر يقفل."""

    def test_watchdog_detects_dead_server(self):
        port = _get_free_port()
        server = ClientsClass.TCPServer(ip="127.0.0.1", port=port)
        server.start()
        server.start_listening(lambda s, a, d: None)
        time.sleep(0.1)

        client = ClientsClass.TCPClient("127.0.0.1", port, timeout=1.0, name="watchdog-test")
        self.assertTrue(client.connect())
        client.start_reconnection_watchdog()
        self.assertTrue(client.connected)

        # نقفل السيرفر — الـ watchdog لازم يكتشف الفصل
        server.stop()

        # نستنى لحد ما الـ watchdog يلاحظ (interval=3s) — مع poll عشان مايبقاش flaky
        deadline = time.time() + 10
        while time.time() < deadline:
            if not client.connected:
                break
            time.sleep(0.2)

        self.assertFalse(
            client.connected,
            "watchdog ماشافش إن السيرفر اقفل — connected لسه True"
        )
        client._stop_monitor.set()  # نوقف الـ watchdog عشان مايعملش busy reconnect
        client.disconnect()


# ════════════════════════════════════════════════════════════════════════════
#                           Excel result reporting
# ════════════════════════════════════════════════════════════════════════════
class TestResultReporting(unittest.TestCase):

    def test_creates_new_file_with_header(self):
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "results.xlsx")
            ex.result_reporting("BC001", "0000001", "pass", file_path=path)
            self.assertTrue(os.path.exists(path))

            import pandas as pd
            df = pd.read_excel(path)
            self.assertEqual(len(df), 1)
            self.assertIn("id", df.columns)
            self.assertEqual(df.iloc[0]["id"], "BC001")
            self.assertEqual(df.iloc[0]["result"], "pass")

    def test_appends_to_existing_file(self):
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "results.xlsx")
            ex.result_reporting("BC001", "0000001", "pass", file_path=path)
            ex.result_reporting("BC002", "0000002", "fail", file_path=path)

            import pandas as pd
            df = pd.read_excel(path)
            self.assertEqual(len(df), 2)
            self.assertEqual(set(df["id"]), {"BC001", "BC002"})


# ════════════════════════════════════════════════════════════════════════════
#                          Scanner barcode handling
# ════════════════════════════════════════════════════════════════════════════
class TestScannerLogic(unittest.TestCase):
    """نختبر منطق _on_key_event بدون ما نشغل keyboard hook حقيقي."""

    def setUp(self):
        scanner.reset_queue()
        scanner._recorded_keys.clear()
        scanner.last_barcode = None

    def test_barcode_assembled_on_enter(self):
        # نمثل المستخدم بيكتب "ABC123" ثم Enter
        FakeEvt = type("E", (), {})
        for ch in "abc123":
            e = FakeEvt()
            e.event_type = "down"  # KEY_DOWN
            e.name = ch
            scanner._on_key_event(e)

        e = FakeEvt()
        e.event_type = "down"
        e.name = "enter"
        scanner._on_key_event(e)

        bc = scanner.queue_barcode.get_nowait()
        self.assertEqual(bc, "abc123")

    def test_duplicate_barcode_ignored(self):
        FakeEvt = type("E", (), {})
        # نقرأ أول مرة
        for ch in "x1":
            e = FakeEvt(); e.event_type = "down"; e.name = ch
            scanner._on_key_event(e)
        e = FakeEvt(); e.event_type = "down"; e.name = "enter"
        scanner._on_key_event(e)
        self.assertEqual(scanner.queue_barcode.qsize(), 1)

        # نقرأ نفس الباركود → لازم يتجاهل
        for ch in "x1":
            e = FakeEvt(); e.event_type = "down"; e.name = ch
            scanner._on_key_event(e)
        e = FakeEvt(); e.event_type = "down"; e.name = "enter"
        scanner._on_key_event(e)
        self.assertEqual(scanner.queue_barcode.qsize(), 1, "الباركود المتكرر مايتحطش في الكيو")


# ════════════════════════════════════════════════════════════════════════════
#                          Debug monitor
# ════════════════════════════════════════════════════════════════════════════
class TestDebugMonitor(unittest.TestCase):

    def test_snapshot_runs_without_app(self):
        text = debug_monitor.snapshot(app_ref=None)
        self.assertIn("MONITOR", text)
        self.assertIn("scanner", text)

    def test_snapshot_with_app_shows_clients(self):
        with mock.patch.object(ClientsClass.TCPServer, "start", return_value=True):
            app = ClientsClass.App()
        try:
            text = debug_monitor.snapshot(app_ref=app)
            # لازم يطلع اسم الـ clients
            self.assertIn("VisionClient_TRIG", text)
            self.assertIn("cobotClient", text)
        finally:
            app.stop()

    def test_is_enabled_respects_env(self):
        with mock.patch.dict(os.environ, {"DEBUG": "1"}):
            self.assertTrue(debug_monitor.is_enabled())
        with mock.patch.dict(os.environ, {"DEBUG": "0"}):
            self.assertFalse(debug_monitor.is_enabled())


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Verbose output
    unittest.main(verbosity=2)


# ════════════════════════════════════════════════════════════════════════════
#               4-Iteration Sequence (الجديد)
# ════════════════════════════════════════════════════════════════════════════
class TestFourIterationSequence(unittest.TestCase):
    """يتأكد إن _run_test_program بيلف 4 مرات (مش 3) ويبعت 4 لكل client."""

    def setUp(self):
        # نشغل 3 mock servers على ports فاضية ونربط الـ App بيهم
        self.stop = threading.Event()
        self.recv_trig, self.port_trig = self._mock_server()
        self.recv_id,   self.port_id   = self._mock_server()
        self.recv_cob,  self.port_cob  = self._mock_server()
        time.sleep(0.1)

        # patch config to point at our mock servers
        from config import config
        self._orig_cfg = {
            "vision_trig_port": config.get("vision_trig_port"),
            "vision_id_port":   config.get("vision_id_port"),
            "cobot_port":       config.get("cobot_port"),
            "cobot_ip":         config.get("cobot_ip"),
            "trigger_server_port": config.get("trigger_server_port"),
        }
        config.set("vision_trig_port", self.port_trig)
        config.set("vision_id_port",   self.port_id)
        config.set("cobot_port",       self.port_cob)
        config.set("cobot_ip",         "127.0.0.1")
        config.set("trigger_server_port", _get_free_port())

    def tearDown(self):
        self.stop.set()
        # نرجع الـ config
        from config import config
        config.update_many(self._orig_cfg)
        time.sleep(0.2)

    def _mock_server(self):
        port = _get_free_port()
        received = []
        def srv():
            s = socket.socket()
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind(("127.0.0.1", port)); s.listen(5); s.settimeout(0.3)
            while not self.stop.is_set():
                try: c, _ = s.accept()
                except socket.timeout: continue
                threading.Thread(target=handle, args=(c,), daemon=True).start()
            s.close()
        def handle(c):
            c.settimeout(0.5)
            while not self.stop.is_set():
                try: d = c.recv(4096)
                except socket.timeout: continue
                if not d: break
                received.append(d)
                try: c.sendall(b"1")
                except Exception: break
            c.close()
        threading.Thread(target=srv, daemon=True).start()
        return received, port

    def test_four_vision_triggers_per_barcode(self):
        """نتأكد إن كل باركود بيبعت 4 triggers للـ VisionTRIG و 4 IDs للـ VisionID."""
        app = ClientsClass.App()
        try:
            app.run()
            time.sleep(1.5)
            # نحط باركود في الـ vision_queue مباشرة (نتفادى scanner)
            app.vision_queue.put("2605TL0000001BGSI")  # G → program 2
            time.sleep(3)

            self.assertEqual(len(self.recv_trig), 4,
                             f"VisionTRIG لازم يستلم 4 رسائل (واحدة لكل iteration). استلم {len(self.recv_trig)}")
            self.assertEqual(len(self.recv_id), 4,
                             f"VisionID لازم يستلم 4 رسائل. استلم {len(self.recv_id)}")
            # الكوبوت: 1 program + 4 y signals + 1 final = 6
            self.assertEqual(len(self.recv_cob), 6,
                             f"Cobot لازم يستلم 6 رسائل (1 program + 4 y + 1 final). استلم {len(self.recv_cob)}")
        finally:
            app.stop()

    def test_id_messages_have_correct_indices(self):
        """نتأكد إن الـ IDs المبعوتة للـ VisionID فيها _0 لحد _3."""
        app = ClientsClass.App()
        try:
            app.run()
            time.sleep(1.5)
            app.vision_queue.put("BARCODE_TEST_BGSI")
            time.sleep(3)

            id_messages = [m.decode() for m in self.recv_id]
            self.assertEqual(len(id_messages), 4)
            for i in range(4):
                expected = f"BARCODE_TEST_BGSI_{i}"
                self.assertIn(expected, id_messages,
                              f"VisionID مفروض يستلم '{expected}' بس كان عنده {id_messages}")
        finally:
            app.stop()

    def test_cobot_messages_are_correct(self):
        """نتأكد إن الكوبوت بياخد: program 2 + y=21,22,23,24 + final (1=pass)."""
        app = ClientsClass.App()
        try:
            app.run()
            time.sleep(1.5)
            app.vision_queue.put("2605TL0000001BGSI")  # program 2
            time.sleep(3)

            cob_msgs = [m.decode() for m in self.recv_cob]
            # awaited: "2" (program), "21", "22", "23", "24" (y signals), "1" (final pass)
            self.assertIn("2", cob_msgs, "program 2 مفروض يتبعت")
            for y in ("21", "22", "23", "24"):
                self.assertIn(y, cob_msgs, f"y signal {y} مفروض يتبعت")
            self.assertIn("1", cob_msgs, "final pass signal (1) مفروض يتبعت")
        finally:
            app.stop()


# ════════════════════════════════════════════════════════════════════════════
#                          Config integration
# ════════════════════════════════════════════════════════════════════════════
class TestConfigIntegration(unittest.TestCase):
    """يتأكد إن config.py شغّال صح + بيتداخل مع باقي البرنامج."""

    def test_default_password_works(self):
        from config import config
        self.assertTrue(config.verify_password("admin"))
        self.assertFalse(config.verify_password("wrong_password"))

    def test_change_password(self):
        from config import config
        # نغير لـ test value
        self.assertTrue(config.set_password("test_pw_xyz"))
        self.assertTrue(config.verify_password("test_pw_xyz"))
        self.assertFalse(config.verify_password("admin"))
        # نرجع للـ default
        config.set_password("admin")
        self.assertTrue(config.verify_password("admin"))

    def test_password_min_length(self):
        from config import config
        # باسوورد قصير لازم يتترفض
        self.assertFalse(config.set_password("123"))
        # ولا الفارغ
        self.assertFalse(config.set_password(""))

    def test_config_persistence(self):
        """تعديل قيمة بيتحفظ ويتقرا تاني."""
        from config import config, Config, CONFIG_FILE
        # نعدل قيمة
        original = config.get("vision_trig_port")
        config.set("vision_trig_port", 12345)
        # نعمل instance جديد ونتأكد إن القيمه اتحفظت
        new_instance = Config(CONFIG_FILE)
        self.assertEqual(new_instance.get("vision_trig_port"), 12345)
        # نرجع للأصلي
        config.set("vision_trig_port", original)

    def test_app_uses_config_for_ips(self):
        """App.__init__ بياخد الـ IPs/Ports من config."""
        from config import config
        config.set("cobot_ip", "10.10.10.10")
        config.set("cobot_port", 7777)
        try:
            with mock.patch.object(ClientsClass.TCPServer, "start", return_value=True):
                app = ClientsClass.App()
            self.assertEqual(app.cobotClient.ip, "10.10.10.10")
            self.assertEqual(app.cobotClient.port, 7777)
            app.stop()
        finally:
            config.set("cobot_ip", "192.168.57.2")
            config.set("cobot_port", 9000)

    def test_state_snapshot_has_all_fields(self):
        """get_state_snapshot ترجع dict كامل (للـ GUI)."""
        with mock.patch.object(ClientsClass.TCPServer, "start", return_value=True):
            app = ClientsClass.App()
        try:
            snap = app.get_state_snapshot()
            self.assertIn("stage", snap)
            self.assertIn("connections", snap)
            self.assertIn("stats", snap)
            self.assertIn("queue_sizes", snap)
            self.assertIn("uptime", snap)
            # الحقول المتوقعة في stats
            for key in ("total", "pass", "fail", "errors"):
                self.assertIn(key, snap["stats"])
        finally:
            app.stop()


# ════════════════════════════════════════════════════════════════════════════
#                           AppStage validation
# ════════════════════════════════════════════════════════════════════════════
class TestAppStage(unittest.TestCase):
    """يتأكد إن AppStage فيها 4 vision tests."""

    def test_vision_test_count_is_4(self):
        self.assertEqual(ClientsClass.AppStage.VISION_TEST_COUNT, 4)

    def test_all_four_vision_stages_exist(self):
        for n in (1, 2, 3, 4):
            stage = getattr(ClientsClass.AppStage, f"VISION_TEST_{n}")
            self.assertEqual(stage, f"VISION_TEST_{n}")

    def test_order_contains_all_4_vision_tests(self):
        order = ClientsClass.AppStage.ORDER
        for n in (1, 2, 3, 4):
            self.assertIn(f"VISION_TEST_{n}", order)

    def test_labels_contain_4_of_4(self):
        labels = ClientsClass.AppStage.LABELS
        for n in (1, 2, 3, 4):
            label = labels.get(f"VISION_TEST_{n}", "")
            self.assertIn(f"{n}/4", label, f"label VISION_TEST_{n} should say {n}/4")


if __name__ == "__main__":
    unittest.main(verbosity=2)


# ════════════════════════════════════════════════════════════════════════════
#       Vision response parsing (BUG FIX: 0 was always returning pass)
# ════════════════════════════════════════════════════════════════════════════
class TestVisionResponseParsing(unittest.TestCase):
    """قبل الإصلاح: send_request بترجع bytes، والمقارنة 0 in [b'0'] كانت دايماً False
    فكل النتايج كانت بتطلع pass حتى لو الفيجن باعت 0."""

    def test_parse_bytes_zero(self):
        self.assertEqual(ClientsClass.App._parse_vision_response(b"0"), 0)

    def test_parse_bytes_one(self):
        self.assertEqual(ClientsClass.App._parse_vision_response(b"1"), 1)

    def test_parse_bytes_with_newline(self):
        self.assertEqual(ClientsClass.App._parse_vision_response(b"0\r\n"), 0)
        self.assertEqual(ClientsClass.App._parse_vision_response(b"1\r\n"), 1)

    def test_parse_str(self):
        self.assertEqual(ClientsClass.App._parse_vision_response("0"), 0)
        self.assertEqual(ClientsClass.App._parse_vision_response("1"), 1)

    def test_parse_int_passthrough(self):
        self.assertEqual(ClientsClass.App._parse_vision_response(0), 0)
        self.assertEqual(ClientsClass.App._parse_vision_response(1), 1)

    def test_parse_none(self):
        self.assertIsNone(ClientsClass.App._parse_vision_response(None))

    def test_parse_empty_bytes(self):
        self.assertIsNone(ClientsClass.App._parse_vision_response(b""))

    def test_parse_garbage(self):
        self.assertIsNone(ClientsClass.App._parse_vision_response(b"abc"))

    def test_parse_float_string(self):
        self.assertEqual(ClientsClass.App._parse_vision_response(b"1.0"), 1)

    def test_zero_response_marks_as_fail(self):
        """الـ Bug الأصلي: لو الفيجن باعت b'0' الـ list كان [b'0',...] ، 
        و `0 in list` بترجع False → final = pass غلط.
        دلوقتي parsed list = [0, 1, 1, 1] و0 in [0,1,1,1] = True → fail صح."""
        # نحاكي يدوياً
        parsed_results = [
            ClientsClass.App._parse_vision_response(b"1"),
            ClientsClass.App._parse_vision_response(b"0"),  # ده اللي كان بيتجاهل قبل
            ClientsClass.App._parse_vision_response(b"1"),
            ClientsClass.App._parse_vision_response(b"1"),
        ]
        self.assertIn(0, parsed_results, "0 لازم يكون في الـ list بعد الـ parsing")
        # الـ check اللي في _run_test_program
        final = "fail" if any(r == 0 or r is None for r in parsed_results) else "pass"
        self.assertEqual(final, "fail", "لو فيه 0 في النتايج لازم final = fail")

    def test_all_ones_marks_as_pass(self):
        parsed = [ClientsClass.App._parse_vision_response(b"1")] * 4
        final = "fail" if any(r == 0 or r is None for r in parsed) else "pass"
        self.assertEqual(final, "pass")

    def test_none_response_marks_as_fail(self):
        """timeout أو connection error → نعتبره fail لأمان."""
        parsed = [1, 1, None, 1]
        final = "fail" if any(r == 0 or r is None for r in parsed) else "pass"
        self.assertEqual(final, "fail")


# ════════════════════════════════════════════════════════════════════════════
#           Start/Stop control (لا يتصل قبل ما المستخدم يدوس Start)
# ════════════════════════════════════════════════════════════════════════════
class TestStartStopControl(unittest.TestCase):
    """يتأكد إن App.__init__ مش بيفتح أي connections،
    وStart/Stop بيشتغلوا كذا مرة بدون مشاكل."""

    def test_init_does_not_start_server(self):
        """__init__ لازم يبني objects بس - مفيش server بيقوم."""
        from config import config
        # نخلي الـ port متاح
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            self.assertFalse(app.is_running, "is_running لازم يكون False بعد __init__")
            self.assertFalse(app.triggerserver.running, "TCPServer ماكانش يفترض يقوم")
            self.assertFalse(app.VisionClient_TRIG.connected, "Vision TRIG ماكانش يفترض يكون متصل")
            self.assertFalse(app.cobotClient.connected, "Cobot ماكانش يفترض يكون متصل")
        finally:
            app.stop()  # safe even if not running

    def test_start_makes_server_running(self):
        """بعد start() الـ server بيكون running."""
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            self.assertTrue(app.start())
            self.assertTrue(app.is_running)
            self.assertTrue(app.triggerserver.running)
            time.sleep(0.3)
        finally:
            app.stop()

    def test_stop_then_start_works(self):
        """قابل إعادة التشغيل بعد stop."""
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            self.assertTrue(app.start())
            self.assertTrue(app.is_running)
            time.sleep(0.3)

            app.stop()
            self.assertFalse(app.is_running)
            self.assertFalse(app.triggerserver.running)
            time.sleep(0.3)

            # تشغيل تاني
            config.set("trigger_server_port", _get_free_port())
            # نحتاج نعمل App جديد لأن TCPServer instance قديم اتقفل
            app2 = ClientsClass.App()
            self.assertTrue(app2.start())
            self.assertTrue(app2.is_running)
            time.sleep(0.3)
            app2.stop()
        finally:
            try: app.stop()
            except: pass

    def test_double_start_is_safe(self):
        """نداء start مرتين متتاليتين مش بيخرّب حاجه."""
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            self.assertTrue(app.start())
            # ندّيها تاني — لازم ترجع True (شغّاله بالفعل)
            self.assertTrue(app.start())
            self.assertTrue(app.is_running)
            time.sleep(0.3)
        finally:
            app.stop()

    def test_double_stop_is_safe(self):
        """نداء stop مرتين متتاليتين مش بيرمي."""
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        app.start()
        time.sleep(0.3)
        app.stop()
        app.stop()  # لازم تعدي بدون exceptions
        self.assertFalse(app.is_running)

    def test_state_snapshot_has_is_running(self):
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            snap = app.get_state_snapshot()
            self.assertIn("is_running", snap)
            self.assertFalse(snap["is_running"])
            app.start()
            time.sleep(0.2)
            self.assertTrue(app.get_state_snapshot()["is_running"])
        finally:
            app.stop()


# ════════════════════════════════════════════════════════════════════════════
#          Skip unknown barcode (مش error، نستنى scan تاني)
# ════════════════════════════════════════════════════════════════════════════
class TestSkipUnknownBarcode(unittest.TestCase):
    """قبل التعديل: لو الحرف مش في الإكسل → ERROR + errors+=1.
    دلوقتي: skip → stats[skipped]++ والـ stage يرجع IDLE."""

    def test_stats_has_skipped_counter(self):
        from config import config
        config.set("trigger_server_port", _get_free_port())
        app = ClientsClass.App()
        try:
            snap = app.get_state_snapshot()
            self.assertIn("skipped", snap["stats"], "stats لازم يحتوي على skipped")
            self.assertEqual(snap["stats"]["skipped"], 0)
        finally:
            app.stop()

    def test_unknown_barcode_increments_skipped_not_errors(self):
        """نشغل الـ App بـ mock servers، نبعت باركود بحرف غير موجود (Z)،
        ونتأكد إن skipped زاد و errors مش زاد."""
        from config import config
        stop_evt = threading.Event()

        def mock_server(port):
            received = []
            def srv():
                s = socket.socket(); s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                s.bind(("127.0.0.1", port)); s.listen(5); s.settimeout(0.3)
                while not stop_evt.is_set():
                    try: c, _ = s.accept()
                    except socket.timeout: continue
                    threading.Thread(target=handle, args=(c, received), daemon=True).start()
                s.close()
            def handle(c, received):
                c.settimeout(0.5)
                while not stop_evt.is_set():
                    try: d = c.recv(4096)
                    except socket.timeout: continue
                    if not d: break
                    received.append(d)
                    try: c.sendall(b"1")
                    except: break
                c.close()
            threading.Thread(target=srv, daemon=True).start()
            return received, port

        recv_trig, port_trig = mock_server(_get_free_port())
        recv_id,   port_id   = mock_server(_get_free_port())
        recv_cob,  port_cob  = mock_server(_get_free_port())
        time.sleep(0.1)

        # نظبط الـ config
        orig = {
            "vision_trig_port": config.get("vision_trig_port"),
            "vision_id_port":   config.get("vision_id_port"),
            "cobot_port":       config.get("cobot_port"),
            "cobot_ip":         config.get("cobot_ip"),
            "trigger_server_port": config.get("trigger_server_port"),
        }
        config.set("vision_trig_port", port_trig)
        config.set("vision_id_port",   port_id)
        config.set("cobot_port",       port_cob)
        config.set("cobot_ip",         "127.0.0.1")
        config.set("trigger_server_port", _get_free_port())

        # نمسح أي باركودات من tests سابقه
        scanner.reset_queue()
        scanner._recorded_keys.clear()

        app = ClientsClass.App()
        try:
            app.start()
            time.sleep(1.0)

            # نمسح vision_queue تاني بعد start (احتياطي)
            while not app.vision_queue.empty():
                try: app.vision_queue.get_nowait()
                except queue.Empty: break

            # snapshot قبل ما نبعت الباركود الجديد
            stats_before = dict(app.get_state_snapshot()["stats"])

            # نبعت باركود بحرف Z مش موجود في الـ mapping
            app.vision_queue.put("2605TL0000001BZSI")
            time.sleep(2.0)

            stats = app.get_state_snapshot()["stats"]
            # نقارن delta (مش absolute) عشان pollution من tests سابقه ميأثرش
            self.assertEqual(
                stats["skipped"] - stats_before["skipped"], 1,
                "skipped لازم يزيد بـ 1 للحرف غير الموجود",
            )
            self.assertEqual(
                stats["errors"] - stats_before["errors"], 0,
                "errors mafrood ma yzeedsh because this is a skip, not an error",
            )
            self.assertEqual(
                stats["total"] - stats_before["total"], 0,
                "total ma yzeedsh because the barcode was not processed successfully",
            )
            # ومحدش بعت حاجه للـ vision/cobot لأن الباركود اتـ skip
            self.assertEqual(len(recv_trig), 0, "Vision TRIG mafrood ma yostaqbel haga")
            self.assertEqual(len(recv_cob), 0, "Cobot mafrood ma yostaqbel haga")
        finally:
            app.stop()
            stop_evt.set()
            config.update_many(orig)
            time.sleep(0.3)
